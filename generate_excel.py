"""
Génération du fichier Excel CBL Coaching avec openpyxl.
Lance: python3 generate_excel.py
Sortie: CBL_Coaching_Alexandre.xlsm (macro-enabled via VBA embarqué)
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from datetime import datetime, timedelta
import os

# ─── Colors ───────────────────────────────────────────────────────────────────
BLACK      = "FF0A0A0A"
DARK       = "FF111111"
DARK2      = "FF1A1A1A"
DARK3      = "FF222222"
BORDER_C   = "FF2A2A2A"
BRAND      = "FF00D4FF"
WARN       = "FFFF9500"
DANGER     = "FFFF3D3D"
SUCCESS    = "FF00D47A"
WHITE      = "FFF0F0F0"
MUTED      = "FF888888"
FAINT      = "FF444444"

def fill(color): return PatternFill("solid", fgColor=color)
def font(bold=False, color=WHITE, sz=10, italic=False):
    return Font(bold=bold, color=color, size=sz, italic=italic, name="Space Grotesk")
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def thin_border(color=BORDER_C):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def bottom_border(color=BORDER_C):
    return Border(bottom=Side(style="thin", color=color))

def style_header_cell(cell, text, color=BRAND, sz=10, bold=True):
    cell.value = text
    cell.fill = fill(DARK2)
    cell.font = font(bold=bold, color=color, sz=sz)
    cell.alignment = align("center")
    cell.border = thin_border()

def style_data_cell(cell, value=None, color=WHITE, bold=False, align_h="center", fill_color=DARK):
    if value is not None:
        cell.value = value
    cell.fill = fill(fill_color)
    cell.font = font(bold=bold, color=color)
    cell.alignment = align(align_h)
    cell.border = thin_border()

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height

# ──────────────────────────────────────────────────────────────────────────────
# 1. Workbook
# ──────────────────────────────────────────────────────────────────────────────
wb = Workbook()

# ──────────────────────────────────────────────────────────────────────────────
# 2. TABLEAU DE BORD
# ──────────────────────────────────────────────────────────────────────────────
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_dash.sheet_properties.tabColor = "00D4FF"
ws_dash.sheet_view.showGridLines = False

# Background
for row in range(1, 60):
    for col in range(1, 20):
        ws_dash.cell(row, col).fill = fill(BLACK)

# ── Title block ──
ws_dash.merge_cells("B2:L2")
c = ws_dash["B2"]
c.value = "CBL COACHING — ALEXANDRE MAFILLE"
c.font = Font(bold=True, color="00D4FF", size=16, name="Space Grotesk")
c.alignment = align("left")
c.fill = fill(BLACK)

ws_dash.merge_cells("B3:L3")
c = ws_dash["B3"]
c.value = "Objectif: Niveau Espoir (Div 3) | Plan: 12 semaines | Baseline: 2026-04-04"
c.font = font(color=MUTED, sz=9)
c.alignment = align("left")
c.fill = fill(BLACK)

# ── Countdown ──
today = datetime.today()
comp_date = datetime(2026, 9, 1)
days_left = (comp_date - today).days
weeks_done = max(0, (today - datetime(2026, 4, 7)).days // 7)

ws_dash.merge_cells("N2:P2")
c = ws_dash["N2"]
c.value = f"{days_left} JOURS"
c.font = Font(bold=True, color="00D4FF", size=18, name="Space Grotesk")
c.alignment = align("center")
c.fill = fill(DARK2)

ws_dash.merge_cells("N3:P3")
c = ws_dash["N3"]
c.value = "avant la compétition"
c.font = font(color=MUTED, sz=9)
c.alignment = align("center")
c.fill = fill(DARK2)

# ── Section: Performances actuelles ──
ws_dash["B5"].value = "PERFORMANCES ACTUELLES → OBJECTIFS 3 MOIS"
ws_dash["B5"].font = font(bold=True, color=BRAND, sz=10)
ws_dash["B5"].fill = fill(BLACK)

headers = ["Mouvement", "Baseline", "Objectif S4", "Objectif S8", "Objectif S12", "Test récent", "Progression", "Statut"]
cols_w  = [18, 12, 12, 12, 12, 14, 14, 10]
for j, (h, w) in enumerate(zip(headers, cols_w), 2):
    cell = ws_dash.cell(6, j)
    style_header_cell(cell, h, color=BRAND if j == 2 else WHITE)
    set_col_width(ws_dash, j, w)

MOVES = [
    ("Tractions",       23, 25, 28, 30),
    ("Muscle-up",       7,  9,  10, 12),
    ("Dips",            30, 35, 38, 40),
    ("Pompes",          52, 58, 62, 65),
    ("Goblet @16kg",    30, 34, 37, 40),
]

for i, (name, base, s4, s8, s12) in enumerate(MOVES, 7):
    row = i
    set_row_height(ws_dash, row, 22)
    bg = DARK if i % 2 == 1 else DARK2

    style_data_cell(ws_dash.cell(row, 2), name, bold=True, align_h="left", fill_color=bg)
    style_data_cell(ws_dash.cell(row, 3), base, fill_color=bg)
    style_data_cell(ws_dash.cell(row, 4), s4, color=MUTED, fill_color=bg)
    style_data_cell(ws_dash.cell(row, 5), s8, color=WARN, fill_color=bg)
    style_data_cell(ws_dash.cell(row, 6), s12, color=BRAND, bold=True, fill_color=bg)

    # Test récent (cell ref vers feuille Progression)
    ws_dash.cell(row, 7).value = f"=IFERROR(MAX(Progression!B{row+2}:Z{row+2}),\"-\")"
    ws_dash.cell(row, 7).fill = fill(bg)
    ws_dash.cell(row, 7).font = font(bold=True, color=SUCCESS)
    ws_dash.cell(row, 7).alignment = align("center")

    # Progression %
    ws_dash.cell(row, 8).value = f"=IFERROR(ROUND((G{row}-C{row})/(G{row}-C{row})*100,0)&\"%\",\"-\")"
    ws_dash.cell(row, 8).fill = fill(bg)
    ws_dash.cell(row, 8).font = font(color=BRAND)
    ws_dash.cell(row, 8).alignment = align("center")

    # Statut (formule conditionnelle)
    ws_dash.cell(row, 9).value = f'=IFERROR(IF(G{row}>={s12},"✓ ATTEINT",IF(G{row}>={s8},"→ EN ROUTE","⚠ À BOOSTER")),"—")'
    ws_dash.cell(row, 9).fill = fill(bg)
    ws_dash.cell(row, 9).font = font()
    ws_dash.cell(row, 9).alignment = align("center")

# ── Section: Semaine courante ──
ws_dash["B13"].value = "SEMAINE EN COURS"
ws_dash["B13"].font = font(bold=True, color=BRAND)
ws_dash["B13"].fill = fill(BLACK)

phase = "Mois 1 — Base Lactique" if weeks_done <= 4 else ("Mois 2 — Spécificité CBL" if weeks_done <= 8 else "Mois 3 — Peaking")
ws_dash.merge_cells("B14:G14")
ws_dash["B14"].value = f"Semaine {weeks_done + 1} / 12  |  {phase}  |  {today.strftime('%d/%m/%Y')}"
ws_dash["B14"].font = font(color=WHITE, sz=11, bold=True)
ws_dash["B14"].fill = fill(DARK2)
ws_dash["B14"].alignment = align("left")
ws_dash.merge_cells("H14:I14")
ws_dash["H14"].value = f"{weeks_done + 1}/12"
ws_dash["H14"].font = Font(bold=True, color="00D4FF", size=18, name="Space Grotesk")
ws_dash["H14"].fill = fill(DARK2)
ws_dash["H14"].alignment = align("center")

for col in range(2, 10):
    ws_dash.cell(14, col).fill = fill(DARK2)

# Column widths
set_col_width(ws_dash, 1, 2)  # A spacer

# Row heights
for r in range(1, 20):
    set_row_height(ws_dash, r, 20)
set_row_height(ws_dash, 2, 32)
set_row_height(ws_dash, 5, 22)
set_row_height(ws_dash, 6, 22)

# ──────────────────────────────────────────────────────────────────────────────
# 3. PROGRESSION DES PERFORMANCES
# ──────────────────────────────────────────────────────────────────────────────
ws_prog = wb.create_sheet("Progression")
ws_prog.sheet_properties.tabColor = "00D47A"
ws_prog.sheet_view.showGridLines = False

for row in range(1, 80):
    for col in range(1, 20):
        ws_prog.cell(row, col).fill = fill(BLACK)

ws_prog.merge_cells("B2:M2")
c = ws_prog["B2"]
c.value = "SUIVI DE PROGRESSION — TESTS DE PERFORMANCE"
c.font = font(bold=True, color=SUCCESS, sz=14)
c.alignment = align("left")
c.fill = fill(BLACK)

# Column headers
perf_headers = ["Date", "Semaine", "Label", "Tractions", "Muscle-up", "Dips", "Pompes", "Goblet @16kg", "Note"]
perf_widths   = [14,    10,        16,     14,           14,           12,    12,       16,              30]

for j, (h, w) in enumerate(zip(perf_headers, perf_widths), 2):
    cell = ws_prog.cell(4, j)
    style_header_cell(cell, h)
    set_col_width(ws_prog, j, w)

# Baseline row
baseline = [datetime(2026, 4, 4), 0, "Baseline", 23, 7, 30, 52, 30, "Point de départ"]
for j, val in enumerate(baseline, 2):
    cell = ws_prog.cell(5, j)
    cell.value = val.strftime("%d/%m/%Y") if isinstance(val, datetime) else val
    cell.fill = fill(DARK2)
    cell.font = font(bold=True, color=BRAND if j in [5,6,7,8,9] else WHITE)
    cell.alignment = align("center" if j != 10 else "left")
    cell.border = thin_border()

# Empty rows for future entries (with borders)
for i in range(6, 56):
    for j in range(2, 11):
        cell = ws_prog.cell(i, j)
        cell.fill = fill(DARK if i % 2 == 0 else DARK2)
        cell.font = font(color=WHITE)
        cell.alignment = align("center" if j != 10 else "left")
        cell.border = thin_border()
    # Date auto-format
    ws_prog.cell(i, 2).number_format = "DD/MM/YYYY"

# Cibles reference block
ref_row = 58
ws_prog.merge_cells(f"B{ref_row}:D{ref_row}")
ws_prog[f"B{ref_row}"].value = "OBJECTIFS DE RÉFÉRENCE"
ws_prog[f"B{ref_row}"].font = font(bold=True, color=BRAND)
ws_prog[f"B{ref_row}"].fill = fill(BLACK)

ref_headers = ["Mouvement", "S4", "S8", "S12 (cible)"]
for j, h in enumerate(ref_headers, 2):
    cell = ws_prog.cell(ref_row + 1, j)
    style_header_cell(cell, h, color=WARN)

ref_data = [
    ("Tractions",    25, 28, 30),
    ("Muscle-up",    9,  10, 12),
    ("Dips",         35, 38, 40),
    ("Pompes",       58, 62, 65),
    ("Goblet @16kg", 34, 37, 40),
]
for i, (name, s4, s8, s12) in enumerate(ref_data, ref_row + 2):
    bg = DARK if i % 2 == 0 else DARK2
    style_data_cell(ws_prog.cell(i, 2), name, align_h="left", fill_color=bg)
    style_data_cell(ws_prog.cell(i, 3), s4, color=MUTED, fill_color=bg)
    style_data_cell(ws_prog.cell(i, 4), s8, color=WARN, fill_color=bg)
    style_data_cell(ws_prog.cell(i, 5), s12, color=BRAND, bold=True, fill_color=bg)

set_col_width(ws_prog, 1, 2)
set_row_height(ws_prog, 2, 30)
set_row_height(ws_prog, 4, 22)

# ──────────────────────────────────────────────────────────────────────────────
# 4. SIMULATION CIRCUIT CBL
# ──────────────────────────────────────────────────────────────────────────────
ws_circ = wb.create_sheet("Circuits CBL")
ws_circ.sheet_properties.tabColor = "FF9500"
ws_circ.sheet_view.showGridLines = False

for row in range(1, 80):
    for col in range(1, 22):
        ws_circ.cell(row, col).fill = fill(BLACK)

ws_circ.merge_cells("B2:N2")
c = ws_circ["B2"]
c.value = "SIMULATION CIRCUITS CBL — SUIVI DES TEMPS"
c.font = font(bold=True, color=WARN, sz=14)
c.alignment = align("left")
c.fill = fill(BLACK)

ws_circ.merge_cells("B3:N3")
c = ws_circ["B3"]
c.value = "Objectif: compléter chaque circuit AVANT le Time Cap. Note tes temps et stations non finies."
c.font = font(color=MUTED, sz=9, italic=True)
c.alignment = align("left")
c.fill = fill(BLACK)

# Circuit table headers
circ_headers = ["Date", "Circuit", "Time Cap", "Temps réalisé", "Stations ✓", "Stations ✗", "Gap vs cap", "% complété", "Note / Ressenti"]
circ_widths   = [14,    30,         12,          16,              14,           14,            14,            14,            28]

for j, (h, w) in enumerate(zip(circ_headers, circ_widths), 2):
    cell = ws_circ.cell(5, j)
    style_header_cell(cell, h, color=WARN)
    set_col_width(ws_circ, j, w)

CIRCUITS = [
    ("OQ7 Amateur 1st Round", "6:00", 8),
    ("OQ7 Amateur 2nd Round", "6:00", 9),
    ("OQ7 Amateur ½ Finale",  "8:00", 10),
    ("OQ7 Amateur Finale",    "10:00", 12),
    ("Premiers Pas 1er Tour", "6:00", 8),
    ("Premiers Pas 2e Tour",  "6:00", 7),
    ("WOD Espoir (Div3)",     "8:00", 9),
]

for i, (name, cap, stations) in enumerate(CIRCUITS, 6):
    bg = DARK if i % 2 == 0 else DARK2
    style_data_cell(ws_circ.cell(i, 2), "", fill_color=bg)  # Date
    ws_circ.cell(i, 2).number_format = "DD/MM/YYYY"
    style_data_cell(ws_circ.cell(i, 3), name, align_h="left", fill_color=bg)
    style_data_cell(ws_circ.cell(i, 4), cap, color=WARN, fill_color=bg)
    style_data_cell(ws_circ.cell(i, 5), "", fill_color=bg)  # Temps réalisé
    style_data_cell(ws_circ.cell(i, 6), "", fill_color=bg)  # Stations ✓
    style_data_cell(ws_circ.cell(i, 7), f"={stations}", fill_color=bg)  # Total stations

    # Gap formule: =IFERROR(E6-D6, "")
    ws_circ.cell(i, 8).value = f"=IFERROR(E{i}-D{i},\"\")"
    ws_circ.cell(i, 8).fill = fill(bg)
    ws_circ.cell(i, 8).font = font(color=DANGER)
    ws_circ.cell(i, 8).alignment = align("center")

    # % complété: =IFERROR(F6/G6*100,0)
    ws_circ.cell(i, 9).value = f"=IFERROR(F{i}/G{i}*100,0)"
    ws_circ.cell(i, 9).fill = fill(bg)
    ws_circ.cell(i, 9).font = font(color=BRAND)
    ws_circ.cell(i, 9).alignment = align("center")
    ws_circ.cell(i, 9).number_format = '0"%"'

    style_data_cell(ws_circ.cell(i, 10), "", align_h="left", fill_color=bg)  # Note

set_col_width(ws_circ, 1, 2)
set_row_height(ws_circ, 2, 30)
set_row_height(ws_circ, 5, 22)

# Circuit detail block
detail_row = 18
ws_circ.merge_cells(f"B{detail_row}:N{detail_row}")
ws_circ[f"B{detail_row}"].value = "DÉTAIL CIRCUITS — RÉFÉRENCE"
ws_circ[f"B{detail_row}"].font = font(bold=True, color=BRAND)
ws_circ[f"B{detail_row}"].fill = fill(BLACK)

circuit_details = {
    "OQ7 Amateur 1st Round — Time Cap 6'": [
        "1 Muscle-up + 10 Push-up", "1 Muscle-up + 15 Push-up", "1 Muscle-up + 20 Push-up",
        "6 Pull-up + 6 Goblet @16kg", "8 Pull-up + 9 Goblet @16kg", "10 Pull-up + 12 Goblet @16kg",
        "8 Pull-up + 9 Goblet @16kg", "6 Pull-up + 6 Goblet @16kg",
    ],
    "OQ7 Amateur 2nd Round — Time Cap 6'": [
        "5 Chest Pull @5kg", "1 Muscle-up", "15 Dips", "12 Dips @10kg", "9 Dips @20kg",
        "12 Pull-up", "9 Pull-up @5kg", "6 Pull-up @10kg", "3 Pull-up @15kg",
    ],
    "WOD Espoir (objectif) — Time Cap 8'": [
        "1 MU + 10 Single Bar Dips + 5 Pull-up", "30 Dips", "15 Squat @40kg",
        "10 Pull-up @5kg", "20 Dips @10kg", "15 Squat @40kg", "10 Dips @20kg",
        "1 MU ou 4 Chest Pull", "15 Pull-up",
    ],
}

cur_row = detail_row + 1
for circuit_name, exercises in circuit_details.items():
    ws_circ.merge_cells(f"B{cur_row}:N{cur_row}")
    c = ws_circ[f"B{cur_row}"]
    c.value = circuit_name
    c.font = font(bold=True, color=WARN, sz=10)
    c.fill = fill(DARK2)
    c.alignment = align("left")
    c.border = thin_border()
    cur_row += 1

    for idx, ex in enumerate(exercises, 1):
        ws_circ.cell(cur_row, 2).value = str(idx)
        ws_circ.cell(cur_row, 2).font = font(bold=True, color=WARN)
        ws_circ.cell(cur_row, 2).fill = fill(DARK)
        ws_circ.cell(cur_row, 2).alignment = align("center")

        ws_circ.merge_cells(f"C{cur_row}:N{cur_row}")
        ws_circ.cell(cur_row, 3).value = ex
        ws_circ.cell(cur_row, 3).font = font(color=WHITE)
        ws_circ.cell(cur_row, 3).fill = fill(DARK)
        ws_circ.cell(cur_row, 3).alignment = align("left")
        cur_row += 1

    cur_row += 1  # blank row

# ──────────────────────────────────────────────────────────────────────────────
# 5. PLAN D'ENTRAÎNEMENT
# ──────────────────────────────────────────────────────────────────────────────
ws_plan = wb.create_sheet("Plan Entraînement")
ws_plan.sheet_properties.tabColor = "FF3D3D"
ws_plan.sheet_view.showGridLines = False

for row in range(1, 100):
    for col in range(1, 20):
        ws_plan.cell(row, col).fill = fill(BLACK)

ws_plan.merge_cells("B2:M2")
c = ws_plan["B2"]
c.value = "PLAN 3 MOIS — CALISTHENICS BATTLE LEAGUE"
c.font = font(bold=True, color=DANGER, sz=14)
c.alignment = align("left")
c.fill = fill(BLACK)

phases = [
    ("MOIS 1 — BASE LACTIQUE",    "Sem 1–4",  "Construire la tolérance lactique et le volume de travail",         BRAND),
    ("MOIS 2 — SPÉCIFICITÉ CBL",  "Sem 5–8",  "Reproduction des circuits CBL sous pression temporelle",           WARN),
    ("MOIS 3 — PEAKING",          "Sem 9–12", "Simulation compétition + décharge contrôlée (deload Sem 11-12)",   DANGER),
]

plan_row = 4
for phase_name, weeks, desc, color in phases:
    ws_plan.merge_cells(f"B{plan_row}:C{plan_row}")
    ws_plan[f"B{plan_row}"].value = phase_name
    ws_plan[f"B{plan_row}"].font = font(bold=True, color=color, sz=11)
    ws_plan[f"B{plan_row}"].fill = fill(DARK2)
    ws_plan[f"B{plan_row}"].alignment = align("left")

    ws_plan[f"D{plan_row}"].value = weeks
    ws_plan[f"D{plan_row}"].font = font(color=MUTED)
    ws_plan[f"D{plan_row}"].fill = fill(DARK2)
    ws_plan[f"D{plan_row}"].alignment = align("center")

    ws_plan.merge_cells(f"E{plan_row}:M{plan_row}")
    ws_plan[f"E{plan_row}"].value = desc
    ws_plan[f"E{plan_row}"].font = font(color=WHITE, italic=True)
    ws_plan[f"E{plan_row}"].fill = fill(DARK2)
    ws_plan[f"E{plan_row}"].alignment = align("left")
    set_row_height(ws_plan, plan_row, 24)
    plan_row += 1

plan_row += 1

# Week-by-week table
week_headers = ["Sem", "Phase", "Séance 1", "Séance 2", "Séance 3", "Séance 4", "Séance 5", "Déload?", "Focus principal"]
week_widths   = [8,    16,      20,         20,          20,          20,          20,          10,        30]

for j, (h, w) in enumerate(zip(week_headers, week_widths), 2):
    cell = ws_plan.cell(plan_row, j)
    style_header_cell(cell, h, color=DANGER)
    set_col_width(ws_plan, j, w)

set_row_height(ws_plan, plan_row, 22)
plan_row += 1

WEEK_PLANS = [
    (1,  "Base Lactique", "Force Pull/MU",    "Lactique Push",   "Circuit x4",      "Récup Active",    "",               False, "Qualité d'exécution"),
    (2,  "Base Lactique", "Force Pull/MU",    "Lactique Push",   "Circuit x4",      "Récup Active",    "",               False, "Récupération entre séries"),
    (3,  "Base Lactique", "Force Pull/MU",    "Lactique Push",   "Circuit x5",      "Récup Active",    "AMRAP 6min",     False, "Tolérance lactique"),
    (4,  "Base Lactique", "Force Pull/MU",    "Lactique Push",   "Circuit x5",      "Simulation légère","Récup Active",  False, "Montée volume"),
    (5,  "Spécificité",   "Force Max Pull",   "Circuit CBL x2",  "Récup Active",    "Densité max",     "Simulation 6min",False, "Premier circuit CBL complet"),
    (6,  "Spécificité",   "Force Max Pull",   "Circuit CBL x2",  "Récup Active",    "Densité max",     "Simulation 6min",False, "Réduire pauses"),
    (7,  "Spécificité",   "Force lestée",     "Circuit CBL x3",  "Récup Active",    "Densité max",     "Simulation x2", False, "Circuit sans pause"),
    (8,  "Spécificité",   "Force lestée",     "Circuit CBL x3",  "Récup Active",    "Intensité courte","Simulation x2", False, "Set 1 + Set 2 enchaînés"),
    (9,  "Peaking",       "Force Activation", "Simulation comp.", "Polissage tech.", "Intensité courte","Récup mentale", False, "Conditions réelles"),
    (10, "Peaking",       "Force Activation", "Simulation comp.", "Polissage tech.", "Intensité courte","Récup mentale", False, "Pic de forme"),
    (11, "Taper",         "Entretien léger",  "Simulation 70%",   "Récup totale",   "",                "",              True,  "Décharger la fatigue"),
    (12, "Taper",         "Activation légère","Feeling circuit",  "Récup totale",   "",                "",              True,  "Arriver frais"),
]

PHASE_COLORS_MAP = {"Base Lactique": BRAND, "Spécificité": WARN, "Peaking": DANGER, "Taper": SUCCESS}

for i, (week, phase, s1, s2, s3, s4, s5, deload, focus) in enumerate(WEEK_PLANS):
    bg = DARK if i % 2 == 0 else DARK2
    row = plan_row + i
    set_row_height(ws_plan, row, 20)

    phase_color = PHASE_COLORS_MAP.get(phase, WHITE)

    style_data_cell(ws_plan.cell(row, 2), week, bold=True, fill_color=bg)
    style_data_cell(ws_plan.cell(row, 3), phase, color=phase_color, fill_color=bg)
    style_data_cell(ws_plan.cell(row, 4), s1, align_h="left", fill_color=bg)
    style_data_cell(ws_plan.cell(row, 5), s2, align_h="left", fill_color=bg)
    style_data_cell(ws_plan.cell(row, 6), s3, align_h="left", fill_color=bg)
    style_data_cell(ws_plan.cell(row, 7), s4, align_h="left", fill_color=bg)
    style_data_cell(ws_plan.cell(row, 8), s5, align_h="left", fill_color=bg)
    deload_color = SUCCESS if deload else FAINT
    style_data_cell(ws_plan.cell(row, 9), "OUI" if deload else "-", color=deload_color, fill_color=bg)
    style_data_cell(ws_plan.cell(row, 10), focus, align_h="left", color=MUTED, fill_color=bg)

set_col_width(ws_plan, 1, 2)
set_row_height(ws_plan, 2, 30)

# ──────────────────────────────────────────────────────────────────────────────
# 6. SAISIE SÉANCES
# ──────────────────────────────────────────────────────────────────────────────
ws_log = wb.create_sheet("Log Séances")
ws_log.sheet_properties.tabColor = "888888"
ws_log.sheet_view.showGridLines = False

for row in range(1, 150):
    for col in range(1, 16):
        ws_log.cell(row, col).fill = fill(BLACK)

ws_log.merge_cells("B2:M2")
c = ws_log["B2"]
c.value = "LOG DES SÉANCES — SAISIE RAPIDE"
c.font = font(bold=True, color=WHITE, sz=14)
c.alignment = align("left")
c.fill = fill(BLACK)

log_headers = ["Date", "Sem", "Type", "Nom séance", "Durée (min)", "RPE /10", "Tractions", "MU", "Dips", "Pompes", "Goblet", "Notes / Circuit"]
log_widths   = [14,    8,     14,     24,           14,            10,         12,           8,    10,     10,       10,       30]

for j, (h, w) in enumerate(zip(log_headers, log_widths), 2):
    cell = ws_log.cell(4, j)
    style_header_cell(cell, h, color=MUTED)
    set_col_width(ws_log, j, w)

for i in range(5, 105):
    bg = DARK if i % 2 == 0 else DARK2
    for j in range(2, 14):
        cell = ws_log.cell(i, j)
        cell.fill = fill(bg)
        cell.font = font(color=WHITE)
        cell.alignment = align("center" if j != 5 else "left")
        cell.border = thin_border()
    ws_log.cell(i, 2).number_format = "DD/MM/YYYY"

set_col_width(ws_log, 1, 2)
set_row_height(ws_log, 2, 30)
set_row_height(ws_log, 4, 22)

# ──────────────────────────────────────────────────────────────────────────────
# 7. PRÉPARATION COMPÉTITION
# ──────────────────────────────────────────────────────────────────────────────
ws_comp = wb.create_sheet("Prépa Compétition")
ws_comp.sheet_properties.tabColor = "FF3D3D"
ws_comp.sheet_view.showGridLines = False

for row in range(1, 60):
    for col in range(1, 18):
        ws_comp.cell(row, col).fill = fill(BLACK)

ws_comp.merge_cells("B2:M2")
c = ws_comp["B2"]
c.value = "PRÉPARATION COMPÉTITION CBL"
c.font = font(bold=True, color=DANGER, sz=14)
c.alignment = align("left")
c.fill = fill(BLACK)

# Countdown
ws_comp["B4"].value = "DATE COMPÉTITION:"
ws_comp["B4"].font = font(bold=True, color=WHITE)
ws_comp["B4"].fill = fill(BLACK)
ws_comp["C4"].value = datetime(2026, 9, 1)
ws_comp["C4"].number_format = "DD/MM/YYYY"
ws_comp["C4"].font = font(bold=True, color=DANGER)
ws_comp["C4"].fill = fill(DARK2)
ws_comp["C4"].alignment = align("center")

ws_comp["D4"].value = "JOURS RESTANTS:"
ws_comp["D4"].font = font(bold=True, color=WHITE)
ws_comp["D4"].fill = fill(BLACK)
ws_comp["E4"].value = f'=DATEDIF(TODAY(),C4,"D")'
ws_comp["E4"].font = Font(bold=True, color="FF3D3D", size=18, name="Space Grotesk")
ws_comp["E4"].fill = fill(DARK2)
ws_comp["E4"].alignment = align("center")
set_row_height(ws_comp, 4, 30)

# Checklist
checklist_title = 6
ws_comp.merge_cells(f"B{checklist_title}:M{checklist_title}")
ws_comp[f"B{checklist_title}"].value = "CHECKLIST PRÉPARATION"
ws_comp[f"B{checklist_title}"].font = font(bold=True, color=WARN)
ws_comp[f"B{checklist_title}"].fill = fill(BLACK)

checklist_headers = ["#", "Élément", "Statut", "Délai", "Notes"]
checklist_widths   = [6,   40,         14,       14,       30]
for j, (h, w) in enumerate(zip(checklist_headers, checklist_widths), 2):
    cell = ws_comp.cell(7, j)
    style_header_cell(cell, h, color=WARN)
    set_col_width(ws_comp, j, w)

checklist_items = [
    ("Physique",   "Compléter le plan 12 semaines",              "En cours",  "S12",         ""),
    ("Physique",   "Tester circuit CBL complet en <6 min",       "À faire",   "S10",         "OQ7 Amateur R1"),
    ("Physique",   "Simulation 2 sets enchaînés",                "À faire",   "S10",         ""),
    ("Physique",   "Test maxi tractions (objectif: 30)",         "À faire",   "S8 + S12",    ""),
    ("Physique",   "Test maxi muscle-up (objectif: 12)",         "À faire",   "S8 + S12",    ""),
    ("Logistique", "Vérifier inscription CBL",                   "À faire",   "8 sem avant", ""),
    ("Logistique", "Confirmer les divisions et circuits attendus","À faire",   "4 sem avant", ""),
    ("Logistique", "Transport + hébergement",                    "À faire",   "6 sem avant", ""),
    ("Matériel",   "Tenu de compétition",                        "À faire",   "4 sem avant", ""),
    ("Matériel",   "Magnésie + grip support",                    "À faire",   "J-7",         ""),
    ("Mental",     "Visualisation circuit CBL x10",              "En cours",  "Continu",     "5 min/jour"),
    ("Mental",     "Protocole warm-up jour J",                   "À faire",   "S10",         ""),
    ("Récup",      "Bilan kiné / physio si besoin",              "À faire",   "S10",         ""),
    ("Récup",      "Protocole sommeil la semaine avant",         "À faire",   "J-7",         ""),
]

STATUS_COLORS = {"En cours": WARN, "À faire": MUTED, "Fait ✓": SUCCESS}

for i, (cat, item, status, delay, notes) in enumerate(checklist_items, 8):
    bg = DARK if i % 2 == 0 else DARK2
    set_row_height(ws_comp, i, 20)

    style_data_cell(ws_comp.cell(i, 2), i - 7, bold=True, fill_color=bg)
    style_data_cell(ws_comp.cell(i, 3), f"[{cat}] {item}", align_h="left", fill_color=bg)

    status_color = STATUS_COLORS.get(status, MUTED)
    style_data_cell(ws_comp.cell(i, 4), status, color=status_color, bold=True, fill_color=bg)
    style_data_cell(ws_comp.cell(i, 5), delay, color=MUTED, fill_color=bg)
    style_data_cell(ws_comp.cell(i, 6), notes, align_h="left", color=MUTED, fill_color=bg)

set_col_width(ws_comp, 1, 2)
set_row_height(ws_comp, 2, 30)
set_row_height(ws_comp, 6, 22)
set_row_height(ws_comp, 7, 22)

# ──────────────────────────────────────────────────────────────────────────────
# 8. Save
# ──────────────────────────────────────────────────────────────────────────────
output = os.path.join(os.path.dirname(__file__), "CBL_Coaching_Alexandre.xlsx")
wb.save(output)
print(f"✅ Fichier généré: {output}")
print("Feuilles créées:")
for s in wb.sheetnames:
    print(f"  · {s}")
